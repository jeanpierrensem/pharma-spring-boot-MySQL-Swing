#!/usr/bin/env python3
"""
Lit « Prix de vente et Achats Avril 2026.xls » (feuilles Sage + Feuil2),
fusionne sans doublon sur la Référence (règle PA×PV≠0 lorsque les deux sources
diffèrent), restructure les colonnes et écrit un .xlsx.

Usage:
  PYTHONPATH=/path/to/xlrd:/path/to/openpyxl python3 excel_restructure_prix_avril.py \
      [--src CHEMIN.xls] [--out CHEMIN.xlsx]
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any, Dict, Optional, Tuple


def normalize_ref(s: str) -> str:
    return str(s).strip()


def _money(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return None


def _pa_pv_non_zero(pa_raw: Any, pv_raw: Any) -> bool:
    """Les deux prix sont présents et strictement différents de zéro."""
    pa = _money(pa_raw)
    pv = _money(pv_raw)
    if pa is None or pv is None:
        return False
    return abs(pa) >= 1e-12 and abs(pv) >= 1e-12


def merge_rows(bk_path: Path) -> Dict[str, Dict[str, Any]]:
    """
    Charge Sage et Feuil2. En cas de doublon sur la référence :
      - on garde la ligne dont le prix d'achat ET le prix de vente sont tous deux non nuls ;
      - si une seule des deux sources remplit cette condition, c'est elle qui est gardée ;
      - si les deux ou aucune des deux la remplissent : on privilégie Sage (priorité fichier source « Dernier prix d'achat »).
    Une même référence répétée dans une même feuille : la dernière occurrence annule les précédentes.
    """

    import xlrd

    bk = xlrd.open_workbook(str(bk_path), formatting_info=False)
    sage = bk.sheet_by_name("Sage")
    feuil2 = bk.sheet_by_name("Feuil2")

    def row_from_sheet(sh: Any, r: int, source: str) -> Dict[str, Any]:
        return {
            "designation": str(sh.cell_value(r, 1) or "").strip(),
            "famille": str(sh.cell_value(r, 2) or "").strip(),
            "stock": sh.cell_value(r, 3),
            "pa": sh.cell_value(r, 4),
            "pv": sh.cell_value(r, 5),
            "from": source,
        }

    sage_by: Dict[str, Dict[str, Any]] = {}
    for r in range(1, sage.nrows):
        ref = normalize_ref(sage.cell_value(r, 0))
        if ref:
            sage_by[ref] = row_from_sheet(sage, r, "Sage")

    feuil_by: Dict[str, Dict[str, Any]] = {}
    for r in range(1, feuil2.nrows):
        ref = normalize_ref(feuil2.cell_value(r, 0))
        if ref:
            feuil_by[ref] = row_from_sheet(feuil2, r, "Feuil2")

    merged: Dict[str, Dict[str, Any]] = {}
    for ref in sorted(set(sage_by) | set(feuil_by)):
        s_row = sage_by.get(ref)
        f_row = feuil_by.get(ref)
        if s_row is None:
            merged[ref] = f_row  # type: ignore[assignment]
        elif f_row is None:
            merged[ref] = s_row
        else:
            s_ok = _pa_pv_non_zero(s_row["pa"], s_row["pv"])
            f_ok = _pa_pv_non_zero(f_row["pa"], f_row["pv"])
            if s_ok ^ f_ok:
                merged[ref] = s_row if s_ok else f_row
            else:
                merged[ref] = s_row

    return merged


# --- Parsing désignation (heuristique) ---

_COND_B = re.compile(r"(?i)\bB\s*/\s*\d+(?:[.,]\d+)?\s*[iIℹ]?\b")
_COND_BL = re.compile(r"(?i)\bB\s*/\s*\d+(?:[,.]\d+)?\s*L\b")


def extract_conditionnement(text: str) -> Tuple[str, str]:
    pieces: list[str] = []
    t = text
    for rx in (_COND_BL, _COND_B):
        while True:
            m = rx.search(t)
            if not m:
                break
            piece = " ".join(m.group(0).split())
            pieces.append(piece.upper().replace("B /", "B/"))
            t = t[: m.start()] + " " + t[m.end() :]
    uniq: list[str] = []
    for p in pieces:
        if p and p not in uniq:
            uniq.append(p)
    return ", ".join(uniq), " ".join(t.split())


_forme_patterns: list[tuple[re.Pattern, str]] = [
    (re.compile(r"(?i)\b(si_iv|injectable|\binj\.?|injection)\b"), "Injection"),
    (re.compile(r"(?i)\bsirop\b"), "Sirop"),
    (re.compile(r"(?i)\bgél(?:ules?)?\b"), "Gélule"),
    (re.compile(r"(?i)\bcomprim(?:és|é|e)?\b"), "Comprimé"),
    (re.compile(r"(?i)\bsachet\b"), "Sachet"),
    (re.compile(r"(?i)\bmp\.?\b(?:\s*cutané[e]?)?|\bliquide\b"), "Liquide externe"),
    (re.compile(r"(?i)\bsolution\b(?!\s+pour\b)"), "Solution"),
    (re.compile(r"(?i)\bsuspension\b"), "Suspension"),
    (re.compile(r"(?i)\bppi\b|\b(vial)\b|\bperfusion\b|\bdrip\b"), "Injectable"),
    (re.compile(r"(?i)\bpommade\b|\bcreme\b|\bcrème\b"), "Cutanée"),
]


def guess_forme(text: str) -> Tuple[Optional[str], str]:
    for rx, label in _forme_patterns:
        if rx.search(text):
            return label, rx.sub(" ", text)
    if re.search(r"(?i)\baiguille\b", text):
        return "Matériel", text
    if re.search(r"(?i)\balcool\b", text):
        return "Liquide", text
    return None, text


# Dosages / concentrations (plusieurs motifs)
_dose_parts = re.compile(
    r"(?<![\w])(?:"
    # mg/ml (sans espace) avant les motifs mg « mot entier »
    r"\d+(?:[,.]\d+)?\s*mg/ml\b"
    r"|(?:\d+(?:[,.]\d+)?(?:\s*\+\s*\d+(?:[,.]\d+)?)*\s*%|\d+(?:[,.]\d+)?\s*°)"
    r"|\d+(?:[,.]\d+)?\s*(?:µg|mcg)\s*(?:/\s*\d+(?:[,.]\d+)?\s*ml\b)?"
    r"|\d+(?:[,.]\d+)?\s*UI\b(?:\s*/\s*ml\b)?"
    r"|\d+(?:[,.]\d+)?\s*IU\b(?:\s*/\s*ml\b)?"
    r"|\d+(?:[,.]\d+)?\s*G\b(?:\s*/\s*\d+(?:[,.]\d+)?\s*mg\b)?"
    r"|\d+(?:[,.]\d+)?\s*mg\b(?:\s*/\s*\d+(?:[,.]\d+)?\s*(?:ml|MG|mg)\b)?(?:\s*\+\s*\d+(?:[,.]\d+)?\s*mg\b)?"
    r"|\d+(?:[,.]\d+)?\s*mg\s*/\s*ml\b"
    r"|(?:\d+(?:[,.]\d+)?\s*(?:µg|µL|µl|µm|µM)\b(?:\s*[/:]\s*[^;,]+)?)"
    r"|\d+(?:[,.]\d+)?\s*ml\b"
    r"|\d+(?:[,.]\d+)?\s*[lL]\b(?![a-zàâäéèêëïîôùûü])"
    r")\b"
)


def _normalize_dose(s: str) -> str:
    return re.sub(r"\s+", "", s.lower())


def dedupe_dose_list(found: list[str]) -> list[str]:
    if len(found) < 2:
        return found
    norm = [_normalize_dose(x) for x in found]
    keep: list[str] = []
    for i, f in enumerate(found):
        nf = norm[i]
        if any(j != i and nf in norm[j] and len(norm[j]) > len(nf) for j in range(len(found))):
            continue
        if f not in keep and all(_normalize_dose(k) != nf for k in keep):
            keep.append(f)
    return keep


def extract_dosages(text: str) -> Tuple[str, str]:
    found: list[str] = []
    t = text
    while True:
        m = _dose_parts.search(t)
        if not m:
            break
        chunk = " ".join(m.group(0).split())
        if chunk:
            found.append(chunk)
        t = t[: m.start()] + " " + t[m.end() :]
    t = " ".join(t.split())
    found = dedupe_dose_list(found)
    return "; ".join(found), t


_parens_trail = re.compile(r"\([^)]*\)\s*$")


def nominal_name_from_remainder(rem: str) -> str:
    s = rem
    s = re.sub(r"[,:;]\s*$", "", s)
    s = _parens_trail.sub("", s)
    s = re.sub(r"(?i)\s*/\s*ml\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.strip(" ,;-–—")


def parse_designation(full: str) -> Dict[str, str]:
    raw = " ".join((full or "").replace("\n", " ").split())
    if not raw:
        return {"nom": "", "dosage": "", "forme": "", "conditionnement": ""}

    cond, t1 = extract_conditionnement(raw)
    doses, t2 = extract_dosages(t1)
    forme_label, t3 = guess_forme(t2)
    forme_str = forme_label or ""

    nom = nominal_name_from_remainder(t3)
    low = nom.lower()
    for token in [" et ", ",", " sous ", " poudre", " blister"]:
        idx = low.find(token)
        if idx > 35:
            nom = nominal_name_from_remainder(nom[:idx])
            break

    return {
        "nom": nom[:500],
        "dosage": doses[:300],
        "forme": forme_str[:120],
        "conditionnement": cond[:120],
    }


def to_float(v: Any) -> Optional[float]:
    return _money(v)


def marge_markup(pa: Optional[float], pv: Optional[float]) -> Optional[float]:
    if pa is None or pv is None or pa <= 0:
        return None
    return (pv - pa) / pa


def write_xlsx(rows: list[Dict[str, Any]], out: Path, date_maj: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"

    headers = [
        "Référence",
        "Nom produit",
        "Dosage",
        "Forme",
        "Conditionnement",
        "Famille",
        "Stock disponible",
        "Seuil minimum",
        "Dernier prix d'achat",
        "Prix de vente",
        "Marge",
        "Fournisseur",
        "Date MAJ",
    ]
    ws.append(headers)
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).font = Font(bold=True)

    for row in rows:
        ws.append(
            [
                row["ref"],
                row["nom"],
                row["dosage"],
                row["forme"],
                row["conditionnement"],
                row["famille"],
                row["stock"],
                row["seuil"],
                row["pa"],
                row["pv"],
                row["marge"],
                row["fournisseur"],
                date_maj,
            ]
        )

    num_fmt_money = '# ##0.00'
    num_fmt_pct = '0.0%'

    pa_col = headers.index("Dernier prix d'achat") + 1
    pv_col = headers.index("Prix de vente") + 1
    marge_col = headers.index("Marge") + 1
    stock_col = headers.index("Stock disponible") + 1

    for r in range(2, ws.max_row + 1):
        ws.cell(r, pa_col).number_format = num_fmt_money
        ws.cell(r, pv_col).number_format = num_fmt_money
        mv = ws.cell(r, marge_col).value
        if mv not in ("", None):
            ws.cell(r, marge_col).number_format = num_fmt_pct
        sc = ws.cell(r, stock_col).value
        if isinstance(sc, (int, float)) and sc != "":
            ws.cell(r, stock_col).number_format = '# ##0'

    for idx, col_name in enumerate(headers, start=1):
        if col_name in ("Stock disponible", "Seuil minimum"):
            ll = len(col_name) + 2
        elif col_name == "Référence":
            ll = 18
        elif col_name in ("Nom produit", "Dosage"):
            ll = 42
        else:
            ll = 14
        ws.column_dimensions[get_column_letter(idx)].width = min(ll, 50)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))


def main() -> None:
    parser = argparse.ArgumentParser()
    default_src = Path("/Users/nsemjean/Desktop/Prix de vente et Achats Avril 2026.xls")
    default_out = Path(
        "/Users/nsemjean/Desktop/Prix vente achats Avril 2026 - restructuré.xlsx"
    )
    parser.add_argument("--src", type=Path, default=default_src)
    parser.add_argument("--out", type=Path, default=default_out)
    parser.add_argument(
        "--date-maj",
        default="Avril 2026",
        help="Libellé date de mise à jour (ex: 30/04/2026 ou Avril 2026)",
    )
    args = parser.parse_args()

    merged = merge_rows(args.src)
    refs_sorted = sorted(merged.keys())

    rows: list[Dict[str, Any]] = []
    for ref in refs_sorted:
        rec = merged[ref]
        parts = parse_designation(rec["designation"])
        pa = to_float(rec["pa"])
        pv = to_float(rec["pv"])
        stock = rec["stock"]
        if isinstance(stock, float) and abs(stock - int(stock)) < 1e-9:
            stock = int(stock)

        markup = marge_markup(pa, pv)

        rows.append(
            {
                "ref": ref,
                "nom": parts["nom"],
                "dosage": parts["dosage"],
                "forme": parts["forme"],
                "conditionnement": parts["conditionnement"],
                "famille": rec["famille"],
                "stock": stock,
                "seuil": "",
                "pa": pa if pa is not None else "",
                "pv": pv if pv is not None else "",
                "marge": markup if markup is not None else "",
                "fournisseur": "",
                "source_sheet": rec.get("from"),
            }
        )

    write_xlsx(rows, args.out, args.date_maj)
    print(f"Écrit : {args.out}")
    print(f"Lignes (hors titre) : {len(rows)}")
    print("Note : Dosage / Forme / Conditionnement sont déduits par heuristique depuis la désignation.")


if __name__ == "__main__":
    main()
